#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys, csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Series, Reference
from operator import itemgetter

reload(sys)
sys.setdefaultencoding('utf-8')


"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
This script is my submission to the MAIS 202 Coding challenge. It is coded in python 2.7. It is composed of 3 functions and some running code at the end of the file.
The 3 functions are 'data','average' and 'graph'. These 3 functions are called by the couple lines of code at the end of the script which are executed automatically when
the script is run. I know there are different ways to run this script but how I usually do it and whiwh I know works for sure is juste to run it in Command Lien with
$python2.7 data.py
I choose to use the openpyxl library because it is a library I know rather well and I have coded with it before. I know that for this knid of work, NumPy is more often used but
unfortunately I don't know it enough yet and I was running out of time thus my decision to go for a 'safer' route.
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""


"""
This function is the first one called. It takes the csv file and creates a reader object for it. It selects only the columns which are asked for and then writes only these
columns in a new spreadsheet in an Excel workbook. 
"""
def data(fileToReadFrom, chosenFieldnames, worksheetToWrite):

    #reads the csv file
    csv.register_dialect('semicolons', delimiter=';')
    fieldnamesreader = csv.DictReader(fileToReadFrom) #find only the columns titles with a dictionnary reader
    fieldnameslist = fieldnamesreader.fieldnames
    L=[]

    #loop to add to a list the column letters only of the colums we want (in chosenFieldnames)
    for index in range(len(chosenFieldnames)):
       L.append(get_column_letter(fieldnameslist.index(chosenFieldnames[index]) + 1))

    #for debugging purposes
    '''
    for item, index in enumerate(L):
        print item, index
    print "\n"
    '''

    #create a new reader object
    reader = csv.reader(fileToReadFrom)

    #loop which finds each cell of the colums we wish to copy and then copies these colums in a new spreadsheet ('worksheetToWrite')
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            if column_letter in L:
                columnletter = get_column_letter((L.index(column_letter)) +1)
                if cell in (None, " "):
                    worksheetToWrite["%s%d" % (columnletter, (row_index + 2))].value = None #failsafe in case the info in the cell is not in the excepted form (is Null)
                else:
                    try:
                        worksheetToWrite["%s%d" % (columnletter, (row_index + 2))].value = cell.lower()
                    except:
                        worksheetToWrite["%s%d" % (columnletter, (row_index + 2))].value = cell.replace(u"\u0003","").lower() #other failsafe

    #loop which writes the titles of the copied colums on the first line of the spreadsheet
    for index in range(len(chosenFieldnames)):
        worksheetToWrite.cell(row = 1, column = (index+1), value=(chosenFieldnames[index]))
    


"""
This is the second function called by my code. It computes the averages for each purposes
"""
def average(worksheetToRead):

    purposes = []
    values = []

    #reads each purpose with its value and adds it to the adequate list
    for row_index, index in enumerate(worksheetToRead.iter_rows()):
        purpose = worksheetToRead["%s%d" % ('A', (row_index + 2))].value
        if purpose is None:
            break
        value = float(worksheetToRead["%s%d" % ('B', (row_index + 2))].value)
        #print purpose, value
        if purpose in purposes:
            i = purposes.index(purpose)
            values[i].append(value)
        else:
            purposes.append(purpose)
            a = [value]
            values.append(a)

    averages = []

    #computes the averages of each sublist for each purpose and puts them in the averages list
    for l in values:
        numElements = len(l)
        result = 0
        for element in l:
            result += element
        averages.append(result/numElements)
        
    #print purposes
    #print averages

    return purposes, averages




"""
This is the third and last functio called by my script. Its purpose is entirely to print the averages for each purposes and create the graph in a new spreadsheet.
"""
def graph(purposes, averages, worksheet):

    #writes the titles in the spreadsheet
    worksheet.cell(row = 1, column = 1, value = "purposes")
    worksheet.cell(row = 1, column = 2, value = "avg_rate")

    #writes the purposes with the corresponding averages values in the spreadsheet
    for index, item in enumerate(purposes):
        worksheet.cell(row = (index+2), column = 1, value = item)
        worksheet.cell(row = (index+2), column = 2, value = averages[index])


    #creates the Excel bar graph
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 13
    chart1.height = 20
    chart1.width = 20
    chart1.title = "Bar Chart"
    chart1.y_axis.title = 'Average Rates'
    chart1.x_axis.title = 'Purpose'

    data = Reference(worksheet, min_col=2, max_col=2, min_row=2, max_row=13)
    cats = Reference(worksheet, min_col=1, max_col=1, min_row=2, max_row=13)
    chart1.add_data(data)
    chart1.set_categories(cats)
    chart1.shape = 4
    worksheet.add_chart(chart1, "E4") #adds the graph to the spreadsheet



"""
This is the part of my code which is automatically executed when my script is run.
"""
wb = Workbook() #creates an Openpyxl Workbook
ws = wb.active #with a new spreadsheet

file = open('data.csv', 'rb') #opens the .csv file
chosenFieldnames = ['purpose','int_rate']
data(file, chosenFieldnames, ws) #call to data() function
 
purposes, averages = average(ws) #call to averages() function

ws_chart = wb.create_sheet(title = "chart") #creates a new spreadsheet where the final averages and the grpah will be added to
graph(purposes, averages, ws_chart) #call to graph() function
wb.remove(ws) #removes the original spreadsheet which was used to add all the necessary information copied form the original .csv
wb.save("result.xlsx") #saves the file in Excel format


